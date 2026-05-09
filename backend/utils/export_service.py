import os
import zipfile
import logging
from datetime import datetime, timedelta, timezone
from urllib.parse import urlparse, unquote

from PyPDF2 import PdfMerger
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XlImage
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

COLUMN_DEFINITIONS = {
    'item_name': {'label': '名称', 'width': 25},
    'purchase_platform': {'label': '购买平台', 'width': 18},
    'amount': {'label': '金额', 'width': 14},
    'purchase_date': {'label': '购物日期', 'width': 14},
    'invoice_date': {'label': '开票日期', 'width': 14},
    'uploader': {'label': '上传人', 'width': 12},
    'invoice_tax_number': {'label': '发票税号', 'width': 22},
    'receipt_image': {'label': '购物图片', 'width': 20},
    'invoice_image': {'label': '发票图片', 'width': 20},
}

# Image column keys that require file download + embedding
IMAGE_COLUMNS = {'receipt_image', 'invoice_image'}

# Target pixel size for embedded images in Excel
IMG_TARGET_WIDTH_PX = 160
IMG_TARGET_HEIGHT_PX = 100


class ExportService:
    def __init__(self, app):
        self.app = app
        self.exports_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'exports')
        os.makedirs(self.exports_dir, exist_ok=True)

    def run_export(self, task_id):
        with self.app.app_context():
            from models import db, ExportTask
            try:
                task = db.session.get(ExportTask, task_id)
                if not task:
                    return
                task.status = 'processing'
                task.progress_percent = 5
                task.progress_message = '正在获取数据...'
                db.session.commit()

                records, snapshot_time = self._fetch_data(task.event_id)
                task.data_snapshot_time = snapshot_time
                task.record_count = len(records)
                task.progress_percent = 15
                task.progress_message = f'已获取 {len(records)} 条记录'
                db.session.commit()

                # Verify and log export order for PDF-Excel consistency
                logger.info(f'导出顺序验证: 共 {len(records)} 条记录, 按创建时间排序')
                for i, rec in enumerate(records):
                    logger.debug(f'  [{i}] id={rec["id"]} type={rec["type"]} '
                                 f'has_invoice={rec.get("has_invoice")} '
                                 f'created_at={rec.get("created_at")}')

                if not records:
                    task.status = 'failed'
                    task.error_message = '没有可导出的数据'
                    task.completed_at = datetime.now(timezone.utc)
                    db.session.commit()
                    return

                columns_config = task.columns_config or {}
                selected_columns = columns_config.get('columns', [])
                options = self._normalize_options(columns_config.get('options', {}))

                logger.info(f'导出配置: columns={selected_columns}, options={options}')

                import tempfile
                with tempfile.TemporaryDirectory() as temp_dir:
                    task.progress_percent = 25
                    task.progress_message = '正在生成Excel文件...'
                    db.session.commit()

                    excel_path = self._generate_excel(records, selected_columns, temp_dir, temp_dir)

                    merged_pdf_path = None
                    invoice_files = []
                    if options.get('merged_pdf') or options.get('individual_invoices'):
                        task.progress_percent = 40
                        task.progress_message = '正在处理发票文件...'
                        db.session.commit()
                        invoice_files, merged_pdf_path = self._process_invoices(records, temp_dir, options)

                    receipt_folder = None
                    if options.get('receipt_images'):
                        task.progress_percent = 70
                        task.progress_message = '正在处理购物凭证图片...'
                        db.session.commit()
                        receipt_folder = self._process_receipt_images(records, temp_dir)

                    task.progress_percent = 85
                    task.progress_message = '正在打包ZIP文件...'
                    db.session.commit()

                    from models import Event
                    event = db.session.get(Event, task.event_id)
                    event_name = event.event_name if event else 'export'
                    zip_path = self._create_zip(event_name, excel_path, merged_pdf_path,
                                                invoice_files, receipt_folder, task.task_id)

                    task.file_path = zip_path
                    task.file_size = os.path.getsize(zip_path)
                    task.status = 'completed'
                    task.progress_percent = 100
                    task.progress_message = '导出完成'
                    task.completed_at = datetime.now(timezone.utc)
                    task.expires_at = datetime.now(timezone.utc) + timedelta(minutes=30)
                    db.session.commit()
                    logger.info(f'导出完成: task_id={task_id}, file={zip_path}, size={task.file_size}')

            except Exception as e:
                logger.error(f'导出失败: task_id={task_id}, error={str(e)}', exc_info=True)
                try:
                    task.status = 'failed'
                    task.error_message = str(e)
                    task.completed_at = datetime.now(timezone.utc)
                    db.session.commit()
                except Exception:
                    pass
            finally:
                try:
                    db.session.remove()
                except Exception:
                    pass

    def _normalize_options(self, options):
        """Normalize camelCase frontend keys to snake_case backend keys."""
        mapping = {
            'mergedPdf': 'merged_pdf',
            'individualInvoices': 'individual_invoices',
            'receiptImages': 'receipt_images',
        }
        result = {}
        for k, v in options.items():
            key = mapping.get(k, k)
            result[key] = v
        return result

    def _fetch_data(self, event_id):
        from models import db, PurchaseRecord, Invoice, User

        purchase_records = PurchaseRecord.query.filter_by(event_id=event_id, is_deleted=False).all()
        invoices = Invoice.query.filter_by(event_id=event_id, is_deleted=False).all()

        logger.info(f'获取数据: event_id={event_id}, purchase_records={len(purchase_records)}, invoices={len(invoices)}')

        all_records = []
        max_updated = None

        for record in purchase_records:
            uploader = db.session.get(User, record.uploader_id)
            updated = record.updated_at or record.created_at
            if max_updated is None or (updated and updated > max_updated):
                max_updated = updated

            rec = {
                'type': 'purchase',
                'id': f'P{record.record_id}',
                'item_name': record.item_name or '',
                'purchase_platform': record.purchase_platform or '',
                'amount': float(record.amount) if record.amount else 0,
                'purchase_date': record.purchase_date.strftime('%Y-%m-%d') if record.purchase_date else '',
                'invoice_date': record.invoice_date.strftime('%Y-%m-%d') if record.invoice_date else '',
                'uploader_name': uploader.real_name if uploader else '未知',
                'invoice_tax_number': record.invoice_tax_number or '',
                'has_invoice': record.has_invoice or False,
                'invoice_file_key': record.invoice_file_key or '',
                'invoice_preview_key': record.invoice_preview_key or '',
                'invoice_original_filename': record.invoice_original_filename or '',
                'receipt_image_url': record.receipt_image_url or '',
                'receipt_image_name': record.receipt_image_name or '',
                'created_at': record.created_at.strftime('%Y-%m-%d %H:%M:%S') if record.created_at else '',
                'raw_created_at': record.created_at,
            }
            all_records.append(rec)
            logger.debug(f'  购买记录 {rec["id"]}: receipt_key={rec["receipt_image_url"][:80] if rec["receipt_image_url"] else "空"}, invoice_key={rec["invoice_file_key"][:80] if rec["invoice_file_key"] else "空"}, tax={rec["invoice_tax_number"]}')

        for invoice in invoices:
            uploader = db.session.get(User, invoice.uploader_id)
            updated = invoice.updated_at or invoice.created_at
            if max_updated is None or (updated and updated > max_updated):
                max_updated = updated

            rec = {
                'type': 'invoice',
                'id': f'I{invoice.invoice_id}',
                'item_name': invoice.project_name or '',
                'purchase_platform': '发票上传',
                'amount': float(invoice.amount) if invoice.amount else 0,
                'purchase_date': invoice.invoice_date.strftime('%Y-%m-%d') if invoice.invoice_date else '',
                'invoice_date': invoice.invoice_date.strftime('%Y-%m-%d') if invoice.invoice_date else '',
                'uploader_name': uploader.real_name if uploader else '未知',
                'invoice_tax_number': '',  # Invoice model has no tax_number field
                'has_invoice': True,
                'invoice_file_key': invoice.image_url or '',
                'invoice_original_filename': invoice.file_name or '',
                'receipt_image_url': '',
                'receipt_image_name': '',
                'created_at': invoice.created_at.strftime('%Y-%m-%d %H:%M:%S') if invoice.created_at else '',
                'raw_created_at': invoice.created_at,
            }
            all_records.append(rec)
            logger.debug(f'  发票记录 {rec["id"]}: invoice_key={rec["invoice_file_key"][:80] if rec["invoice_file_key"] else "空"}, tax={rec["invoice_tax_number"]}')

        all_records.sort(key=lambda x: x.get('raw_created_at') or datetime.min.replace(tzinfo=timezone.utc))
        return all_records, max_updated

    def _generate_excel(self, records, selected_columns, output_dir, temp_dir):
        wb = Workbook()
        ws = wb.active
        ws.title = '购物记录'

        header_fill = PatternFill(start_color='667eea', end_color='667eea', fill_type='solid')
        header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
        header_alignment = Alignment(horizontal='center', vertical='center')
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        if not selected_columns:
            raise ValueError('selected_columns is empty')

        has_image_cols = any(c in IMAGE_COLUMNS for c in selected_columns)
        logger.info(f'生成Excel: 列={selected_columns}, 含图片列={has_image_cols}')

        # Write headers
        for col_idx, col_key in enumerate(selected_columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=COLUMN_DEFINITIONS[col_key]['label'])
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = COLUMN_DEFINITIONS[col_key]['width']

        # Image columns need wider columns to accommodate embedded images
        for col_idx, col_key in enumerate(selected_columns, 1):
            if col_key in IMAGE_COLUMNS:
                col_letter = get_column_letter(col_idx)
                ws.column_dimensions[col_letter].width = (IMG_TARGET_WIDTH_PX / 7) + 2

        # Write data rows
        data_font = Font(name='微软雅黑', size=10)
        img_fallback_font = Font(name='微软雅黑', size=8, color='999999')
        cell_alignment = Alignment(vertical='center', wrap_text=True)
        img_success_count = 0
        img_fail_count = 0

        for row_idx, record in enumerate(records, 2):
            # Determine if this row will have embedded images for row height
            row_has_image = False
            for col_key in selected_columns:
                if col_key in IMAGE_COLUMNS:
                    fk = self._get_image_file_key(record, col_key)
                    if fk:
                        row_has_image = True
                        break

            if row_has_image:
                ws.row_dimensions[row_idx].height = IMG_TARGET_HEIGHT_PX * 0.8
            else:
                ws.row_dimensions[row_idx].height = 22

            for col_idx, col_key in enumerate(selected_columns, 1):
                if col_key in IMAGE_COLUMNS:
                    # has_invoice guard for invoice_image column
                    if col_key == 'invoice_image' and not record.get('has_invoice', False):
                        cell = ws.cell(row=row_idx, column=col_idx, value='无发票信息')
                        cell.font = data_font
                        cell.alignment = cell_alignment
                        cell.border = thin_border
                        continue

                    file_key = self._get_image_file_key(record, col_key)
                    if not file_key:
                        label = '无发票信息' if col_key == 'invoice_image' else '无'
                        cell = ws.cell(row=row_idx, column=col_idx, value=label)
                        cell.font = data_font
                        cell.alignment = cell_alignment
                        cell.border = thin_border
                        if col_key == 'invoice_image':
                            img_fail_count += 1
                        continue

                    img_path = self._download_and_save_image(file_key, temp_dir, record, col_key)
                    if img_path and os.path.exists(img_path):
                        try:
                            img = XlImage(img_path)
                            img.width = IMG_TARGET_WIDTH_PX
                            img.height = IMG_TARGET_HEIGHT_PX
                            cell_ref = f'{get_column_letter(col_idx)}{row_idx}'
                            ws.add_image(img, cell_ref)
                            cell = ws.cell(row=row_idx, column=col_idx, value='[图片]')
                            cell.font = img_fallback_font
                            cell.alignment = cell_alignment
                            cell.border = thin_border
                            img_success_count += 1
                        except Exception as e:
                            logger.warning(f'插入图片失败: record={record.get("id")}, col={col_key}, path={img_path}, error={str(e)}')
                            cell = ws.cell(row=row_idx, column=col_idx, value='图片加载失败')
                            cell.font = data_font
                            cell.alignment = cell_alignment
                            cell.border = thin_border
                            img_fail_count += 1
                    else:
                        logger.debug(f'图片下载失败: record={record.get("id")}, col={col_key}, key={file_key[:60] if file_key else "空"}')
                        cell = ws.cell(row=row_idx, column=col_idx, value='资源未找到')
                        cell.font = data_font
                        cell.alignment = cell_alignment
                        cell.border = thin_border
                        img_fail_count += 1
                else:
                    value = self._get_column_value(record, col_key)
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.font = data_font
                    cell.border = thin_border
                    if col_key == 'amount':
                        cell.number_format = '#,##0.00'
                        cell.alignment = Alignment(horizontal='right', vertical='center')
                    else:
                        cell.alignment = cell_alignment

        logger.info(f'Excel图片统计: 成功={img_success_count}, 失败={img_fail_count}')
        excel_path = os.path.join(output_dir, '购物信息.xlsx')
        wb.save(excel_path)
        return excel_path

    def _get_image_file_key(self, record, col_key):
        """Get the storage file key for an image column."""
        if col_key == 'receipt_image':
            return record.get('receipt_image_url', '')
        elif col_key == 'invoice_image':
            # Prefer preview JPG over original PDF for embedding in Excel
            return (record.get('invoice_preview_key', '')
                    or record.get('invoice_file_key', '')
                    or record.get('image_url', ''))
        return ''

    def _download_and_save_image(self, file_key, temp_dir, record, col_key):
        """Download an image file and save to temp dir, return the local path."""
        if not file_key:
            return None

        # Determine filename and extension
        if col_key == 'receipt_image':
            original_name = record.get('receipt_image_name', '')
            prefix = 'receipt'
        else:
            original_name = record.get('invoice_original_filename', '') or record.get('file_name', '')
            prefix = 'invoice'

        # Get extension from file_key first (actual file), then original name, or default
        ext = ''
        if file_key:
            ext = os.path.splitext(file_key.split('?')[0])[1].lower()
        if not ext and original_name:
            ext = os.path.splitext(original_name)[1].lower()
        if not ext:
            ext = '.jpg'  # default

        record_id = record.get('id', 'unknown')
        safe_filename = f'{prefix}_{record_id}{ext}'
        local_path = os.path.join(temp_dir, 'img_temp', safe_filename)

        os.makedirs(os.path.dirname(local_path), exist_ok=True)

        # Download file bytes
        file_bytes = self._download_file(file_key)
        if not file_bytes:
            return None

        logger.info(f'图片下载成功: record={record_id}, col={col_key}, size={len(file_bytes)} bytes')

        with open(local_path, 'wb') as f:
            f.write(file_bytes)

        # For PDF files, convert first page to image
        if ext == '.pdf':
            try:
                import fitz  # PyMuPDF
                pdf_doc = fitz.open(local_path)
                if pdf_doc.page_count > 0:
                    page = pdf_doc[0]
                    pix = page.get_pixmap(matrix=fitz.Matrix(2, 2))
                    img_path = local_path.replace('.pdf', '.png')
                    pix.save(img_path)
                    pdf_doc.close()
                    logger.info(f'PDF转图片成功: {img_path}')
                    return img_path
                pdf_doc.close()
            except Exception as e:
                logger.warning(f'PDF转图片失败: {local_path}, error={str(e)}')
                return None

        # For non-image files, skip
        if ext not in ('.jpg', '.jpeg', '.png', '.gif', '.bmp', '.webp'):
            logger.warning(f'不支持的图片格式: ext={ext}, record={record.get("id")}, file_key={file_key[:80]}')
            return None

        return local_path

    def _get_column_value(self, record, col_key):
        if col_key == 'item_name':
            return record.get('item_name', '')
        elif col_key == 'purchase_platform':
            return record.get('purchase_platform', '')
        elif col_key == 'amount':
            return record.get('amount', 0)
        elif col_key == 'purchase_date':
            return record.get('purchase_date', '')
        elif col_key == 'invoice_date':
            return record.get('invoice_date', '')
        elif col_key == 'uploader':
            return record.get('uploader_name', '')
        elif col_key == 'invoice_tax_number':
            if not record.get('has_invoice', False):
                return '无发票信息'
            tax = record.get('invoice_tax_number', '')
            return tax if tax else '无'
        return ''

    def _extract_cos_key(self, file_key):
        """Extract bare COS key from various formats (presigned URL, full URL, or raw key)."""
        if not file_key:
            return None

        # If it looks like a URL, parse and extract the path
        if file_key.startswith('http://') or file_key.startswith('https://'):
            try:
                parsed = urlparse(file_key)
                # Path is like /invoices/xxx/file.jpg — strip leading /
                path = unquote(parsed.path).lstrip('/')
                if path:
                    return path
            except Exception:
                pass
            return None

        # Already a bare COS key or local path
        return file_key

    def _download_file(self, file_key):
        """Download a file by its key. Handles presigned URLs, COS keys, and local paths."""
        if not file_key:
            return None

        logger.debug(f'尝试下载文件: {file_key[:80]}')

        # Strategy 1: Check local uploads directory first (fastest)
        uploads_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'uploads')

        # Handle /uploads/xxx local paths
        if file_key.startswith('/uploads/'):
            local_path = os.path.join(uploads_dir, file_key.replace('/uploads/', ''))
            if os.path.exists(local_path):
                logger.debug(f'从本地读取: {local_path}')
                with open(local_path, 'rb') as f:
                    return f.read()
            logger.warning(f'本地文件不存在: {local_path}')
            return None

        # Strategy 2: If it's a full URL, download directly via HTTP
        if file_key.startswith('http://') or file_key.startswith('https://'):
            try:
                import requests as requests_lib
                resp = requests_lib.get(file_key, timeout=30)
                resp.raise_for_status()
                logger.debug(f'从URL下载成功: {len(resp.content)} bytes')
                return resp.content
            except Exception as e:
                logger.warning(f'从URL下载失败: {file_key[:80]}, error={str(e)}')
            return None

        # Strategy 3: Use storage backend (local or COS)
        try:
            from utils.storage import storage_manager
            if storage_manager.is_available():
                content = storage_manager.download_file(file_key)
                if content:
                    logger.debug(f'存储下载成功: {len(content)} bytes')
                    return content
        except Exception as e:
            logger.warning(f'存储下载失败: {file_key}, error={str(e)}')

        # Strategy 4: Try as relative path in uploads dir
        local_path = os.path.join(uploads_dir, file_key)
        if os.path.exists(local_path):
            logger.debug(f'从本地相对路径读取: {local_path}')
            with open(local_path, 'rb') as f:
                return f.read()

        logger.warning(f'文件下载失败(所有方式): {file_key[:80]}')
        return None

    def _process_invoices(self, records, temp_dir, options):
        invoice_files = []
        invoices_dir = os.path.join(temp_dir, '发票单独')
        os.makedirs(invoices_dir, exist_ok=True)

        download_count = 0
        skip_count = 0

        for idx, record in enumerate(records):
            if not record.get('has_invoice') or not record.get('invoice_file_key'):
                skip_count += 1
                continue

            file_bytes = self._download_file(record['invoice_file_key'])
            if not file_bytes:
                skip_count += 1
                continue

            download_count += 1
            tax_number = record.get('invoice_tax_number', '')
            if tax_number:
                filename = f'{tax_number}_{idx + 1}.pdf'
            else:
                original = record.get('invoice_original_filename', '')
                if not original:
                    filename = f'发票_{idx + 1}.pdf'
                else:
                    name, ext = os.path.splitext(original)
                    filename = f'{name}.pdf' if ext.lower() != '.pdf' else original

            filepath = os.path.join(invoices_dir, filename)
            with open(filepath, 'wb') as f:
                f.write(file_bytes)
            invoice_files.append(filepath)

        logger.info(f'发票处理: 下载={download_count}, 跳过={skip_count}, 总文件={len(invoice_files)}')

        merged_pdf_path = None
        if options.get('merged_pdf') and invoice_files:
            merger = PdfMerger()
            successful_merges = 0
            for pdf_file in invoice_files:
                try:
                    merger.append(pdf_file)
                    successful_merges += 1
                except Exception as e:
                    logger.warning(f'合并PDF失败: {pdf_file}, error={str(e)}')
            if successful_merges > 0:
                merged_pdf_path = os.path.join(temp_dir, '发票汇总.pdf')
                merger.write(merged_pdf_path)
                merger.close()
                logger.info(f'PDF合并成功: {successful_merges}/{len(invoice_files)} 个文件')
            else:
                merger.close()
                logger.warning('PDF合并失败: 所有文件都无法合并，已跳过生成合并PDF')

        return invoice_files, merged_pdf_path

    def _process_receipt_images(self, records, temp_dir):
        receipt_dir = os.path.join(temp_dir, '购物凭证')
        os.makedirs(receipt_dir, exist_ok=True)

        download_count = 0
        skip_count = 0

        for idx, record in enumerate(records):
            if not record.get('receipt_image_url'):
                skip_count += 1
                continue

            file_bytes = self._download_file(record['receipt_image_url'])
            if not file_bytes:
                skip_count += 1
                continue

            download_count += 1
            original_name = record.get('receipt_image_name', '')
            if original_name:
                ext = os.path.splitext(original_name)[1] or '.jpg'
                safe_name = "".join(c for c in record.get("item_name", "凭证") if c.isalnum() or c in (' ', '-', '_')).strip()
                filename = f'{safe_name}_{idx + 1}{ext}'
            else:
                filename = f'凭证_{idx + 1}.jpg'

            filepath = os.path.join(receipt_dir, filename)
            with open(filepath, 'wb') as f:
                f.write(file_bytes)

        logger.info(f'购物凭证处理: 下载={download_count}, 跳过={skip_count}')
        return receipt_dir

    def _create_zip(self, event_name, excel_path, merged_pdf_path, invoice_files, receipt_folder, task_id):
        safe_name = "".join(c for c in event_name if c.isalnum() or c in (' ', '-', '_')).strip()
        timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
        zip_filename = f'{safe_name}_{task_id}_{timestamp}.zip'
        zip_path = os.path.join(self.exports_dir, zip_filename)

        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zf:
            if excel_path and os.path.exists(excel_path):
                zf.write(excel_path, os.path.basename(excel_path))
                logger.info(f'  ZIP: 添加Excel ({os.path.getsize(excel_path)} bytes)')

            if merged_pdf_path and os.path.exists(merged_pdf_path):
                zf.write(merged_pdf_path, '发票汇总.pdf')
                logger.info(f'  ZIP: 添加发票汇总 ({os.path.getsize(merged_pdf_path)} bytes)')

            for pdf_file in invoice_files:
                if os.path.exists(pdf_file):
                    zf.write(pdf_file, f'发票单独/{os.path.basename(pdf_file)}')
            logger.info(f'  ZIP: 添加 {len(invoice_files)} 个单独发票文件')

            if receipt_folder and os.path.exists(receipt_folder):
                receipt_count = 0
                for root, dirs, files in os.walk(receipt_folder):
                    for file in files:
                        file_path = os.path.join(root, file)
                        arcname = os.path.relpath(file_path, receipt_folder)
                        zf.write(file_path, f'购物凭证/{arcname}')
                        receipt_count += 1
                logger.info(f'  ZIP: 添加 {receipt_count} 个购物凭证文件')
            else:
                logger.info(f'  ZIP: 无购物凭证文件夹 (receipt_folder={receipt_folder})')

        return zip_path

    def check_cache(self, event_id, columns_config):
        from models import db, ExportTask, PurchaseRecord, Invoice

        self._cleanup_expired()

        columns_hash = str(sorted(columns_config.get('columns', [])))
        options = columns_config.get('options', {})

        recent_task = ExportTask.query.filter_by(
            event_id=event_id,
            status='completed'
        ).order_by(ExportTask.created_at.desc()).first()

        if not recent_task:
            return None

        if not recent_task.file_path or not os.path.exists(recent_task.file_path):
            return None

        if recent_task.expires_at and recent_task.expires_at < datetime.now(timezone.utc):
            return None

        stored_config = recent_task.columns_config or {}
        if str(sorted(stored_config.get('columns', []))) != columns_hash:
            return None
        if stored_config.get('options', {}) != options:
            return None

        max_updated = db.session.query(
            db.func.max(PurchaseRecord.updated_at)
        ).filter_by(event_id=event_id, is_deleted=False).scalar()

        max_updated_inv = db.session.query(
            db.func.max(Invoice.updated_at)
        ).filter_by(event_id=event_id, is_deleted=False).scalar()

        latest_update = max(
            x for x in [max_updated, max_updated_inv, recent_task.data_snapshot_time]
            if x is not None
        ) if any(x is not None for x in [max_updated, max_updated_inv, recent_task.data_snapshot_time]) else None

        if latest_update and recent_task.data_snapshot_time and latest_update > recent_task.data_snapshot_time:
            return None

        return recent_task

    def _cleanup_expired(self):
        from models import db, ExportTask

        expired_tasks = ExportTask.query.filter(
            ExportTask.expires_at < datetime.now(timezone.utc),
            ExportTask.status == 'completed'
        ).all()

        for task in expired_tasks:
            if task.file_path and os.path.exists(task.file_path):
                try:
                    os.remove(task.file_path)
                    logger.info(f'已清理过期导出文件: {task.file_path}')
                except Exception as e:
                    logger.warning(f'清理文件失败: {task.file_path}, error={str(e)}')

            db.session.delete(task)

        db.session.commit()
